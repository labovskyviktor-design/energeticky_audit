"""Extract city climate data from Excel spreadsheet."""
import openpyxl
import json

wb = openpyxl.load_workbook(
    r"C:\Users\42191\.gemini\antigravity\scratch\energy-audit\docs\excel\01_Pomôcka - klimatické údaje.xlsx",
    data_only=True,
)

# Sheet 1: Monthly temperatures
ws = wb["Mesačný priemer vonk. teploty"]
cities = []

for start_row in [6, 20, 34, 48]:
    for col_offset in [0, 5, 10, 15]:
        name_col = 2 + col_offset
        temp_col = 4 + col_offset
        city_name = ws.cell(start_row, name_col).value
        if not city_name:
            continue
        oblast = ws.cell(start_row + 1, name_col).value
        altitude = ws.cell(start_row + 2, name_col).value

        temps = {}
        months_roman = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
        for m in range(12):
            temp_val = ws.cell(start_row + 1 + m, temp_col).value
            if temp_val is not None:
                temps[months_roman[m]] = round(float(temp_val), 2)

        # Calculate heating season: months where θe < 13°C
        heating_temps = [t for t in temps.values() if t < 13]
        theta_em = round(sum(heating_temps) / len(heating_temps), 2) if heating_temps else None

        # Heating days: standard STN values by region
        heating_days_map = {1: 212, 2: 222, 3: 232}
        oblast_num = int(oblast) if oblast else 2
        heating_days = heating_days_map.get(oblast_num, 222)

        cities.append({
            "name": str(city_name).strip(),
            "oblast": int(oblast) if oblast else None,
            "altitude": int(altitude) if altitude else None,
            "monthly_temps": temps,
            "theta_e_m": theta_em,
            "heating_days": heating_days,
        })

# Output
out = json.dumps(cities, ensure_ascii=False, indent=2)
with open(r"C:\Users\42191\.gemini\antigravity\scratch\energy-audit\docs\cities.json", "w", encoding="utf-8") as f:
    f.write(out)

for c in cities:
    print(f"{c['name']:25s} | alt={str(c['altitude']):>4s}m | oblast={c['oblast']} | theta_e_m={c['theta_e_m']} | days={c['heating_days']}")

print(f"\nTotal cities: {len(cities)}")
